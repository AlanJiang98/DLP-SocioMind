"""
Author: Jianping Jiang (alanjjp98@gmail.com)

File: evolution.py
Description: Evolution of AI society
"""
import numpy as np
from statsmodels.stats.proportion import proportions_ztest, proportion_confint
import csv
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from scipy import stats

# Wilson confidence interval:

# Wilson, E. B. (1927). Probable inference, the law of succession, and statistical inference. Journal of the American Statistical Association, 22(158), 209-212.
# Newcombe, R. G. (1998). Two-sided confidence intervals for the single proportion: comparison of seven methods. Statistics in medicine, 17(8), 857-872.

# This concept can be related to Item Response Theory (IRT), which is widely used in educational assessment and psychology:
# Embretson, S. E., & Reise, S. P. (2000). Item response theory for psychologists. L. Erlbaum Associates.
def adjusted_accuracy(correct, total, choices):
    guessing_accuracy = 1.0 / choices
    actual_accuracy = correct / total
    adjusted_acc = (actual_accuracy - guessing_accuracy) / (1 - guessing_accuracy)
    return adjusted_acc


is_adjust_acc = False



database_file_path = r"controllability_questionaire_base.csv"
database = {}
id_to_item = {
    0:'name',
    1:'item_id',
    2:'A_answer',
    3:'B_answer',
    4:'A_nums',
    5:'B_nums',
    6:'is_item',
    7:'attr_type',
    8: 'script_type',
    9: 'plot_id',
    12: 'A_method_type',
    13: 'B_method_type',
}
with open(database_file_path, encoding='utf-8') as csvfile:
    csvreader = csv.reader(csvfile, delimiter=',')
    for row in csvreader:
        if row[1].isdigit():
            database[int(row[1])] = {}
            for key, value in id_to_item.items():
                database[int(row[1])][value] = row[key]

database = dict(sorted(database.items(), key=lambda item: item[0]))

questionaire_file_path = r"output\242181177_37_2.xlsx"
wb = load_workbook(filename=questionaire_file_path)
sheet = wb[wb.sheetnames[0]]



questionaire_data = {}
for row in sheet.iter_rows(values_only=True):
    if type(row[0]) is int or row[0].isdigit():
        questionaire_data[int(row[0])] = {}
        for item_id, answer in enumerate(row[13:]):
            questionaire_data[int(row[0])][item_id] = answer
            


item_id_to_database = {}

acc_item_id = -1
for key, value in database.items():
    if value['A_answer'].upper() in ['A', 'B', 'C', 'D', 'E']:
        acc_item_id += 1
        item_id_to_database[acc_item_id] = {'AorB': 'A', 'db_id': key}

    if value['B_answer'].upper() in ['A', 'B', 'C', 'D', 'E']:
        acc_item_id += 1
        item_id_to_database[acc_item_id] = {'AorB': 'B', 'db_id': key}


results = {
    'GA': {},
    'Psycho': {},
}

def to_id(s: str):
    return ord(s.upper()) - ord('A') + 1

for test_id, test_res in questionaire_data.items():
    for t, (item_id, answer) in enumerate(test_res.items()):
        if t >= 101:
            break
        if type(answer) is not int or answer < 0:
            continue
        else:
            AorB = item_id_to_database[item_id]['AorB']
            db_id = item_id_to_database[item_id]['db_id']
            method_type = database[db_id][AorB + '_method_type']
            
            if method_type not in results.keys():
                continue
            
            attr_type = database[db_id]['attr_type']
            
            if attr_type not in results[method_type].keys():
                results[method_type][attr_type] = {}
                
            if db_id not in results[method_type][attr_type].keys():
                results[method_type][attr_type][db_id] = {'corr': 0, 'all': 0}
            
            if answer == to_id(database[db_id][AorB + '_answer']):
                results[method_type][attr_type][db_id]['corr'] += 1
            
            results[method_type][attr_type][db_id]['all'] += 1
            
            results[method_type][attr_type][db_id]['option'] = database[db_id][AorB + '_nums']


controllability = {}
control_final = {}
for method in results.keys():
    if method not in controllability.keys():
        controllability[method] = {}
    for attr in results[method].keys():
        correct_answers = {'3_option': 0, '4_option': 0, '5_option': 0}
        total_questions = {'3_option': 0, '4_option': 0, '5_option': 0}
        guess_probs = {'3_option': 1/3, '4_option': 1/4, '5_option': 1/5}
        for db_id in results[method][attr].keys():
            if results[method][attr][db_id]['option'] == '3':
                correct_answers['3_option'] += results[method][attr][db_id]['corr']
                total_questions['3_option'] += results[method][attr][db_id]['all']
            elif results[method][attr][db_id]['option'] == '4':
                correct_answers['4_option'] += results[method][attr][db_id]['corr']
                total_questions['4_option'] += results[method][attr][db_id]['all']
            elif results[method][attr][db_id]['option'] == '5':
                correct_answers['5_option'] += results[method][attr][db_id]['corr']
                total_questions['5_option'] += results[method][attr][db_id]['all']

        expected_guesses = {key: total_questions[key] * guess_probs[key] for key in total_questions}
        total_expected_correct_guesses = sum(expected_guesses.values())
        total_actual_correct_answers = sum(correct_answers.values())
        
        total_questions_all = sum(total_questions.values())
        if is_adjust_acc:
            accuracy = (total_actual_correct_answers - total_expected_correct_guesses) / \
                (total_questions_all - total_expected_correct_guesses + 1e-6)
        else:
            accuracy = total_actual_correct_answers / total_questions_all
        
        conf_int_low, conf_int_high = proportion_confint(total_actual_correct_answers, total_questions_all,
                                                 alpha=0.05, method='wilson')

        if attr not in controllability[method].keys():
            controllability[method][attr] = {}
        
        controllability[method][attr]['accuracy'] = accuracy
        controllability[method][attr]['conf_int_low'] = conf_int_low
        controllability[method][attr]['conf_int_high'] = conf_int_high
        
        print(f'{method} {attr} {accuracy:.2f} [{conf_int_low:.2f}, {conf_int_high:.2f}]')
        

ablation_file_path = r"ablation.csv"
database_ablation = {}
with open(ablation_file_path, encoding='utf-8') as csvfile:
    csvreader = csv.reader(csvfile, delimiter=',')
    for row in csvreader:
        if row[6] in ['A', 'B', 'C', 'D', 'E']:
            id = int(row[3]) * 5 + ord(row[6]) - ord('A')
            database_ablation[id] = {
                'type': row[0],
                'script': row[1],
                'file_path': row[2],
                'sheet': int(row[3]),
                'column': ord(row[6]) - ord('A'),
            }

ablation = {}
for test_id, test_res in questionaire_data.items():
    for t, (item_id, answer) in enumerate(test_res.items()):
        if t <= 101 or t >= 142:
            continue
        if type(answer) is not int or answer < 0:
            continue
        else:
            ab_id =  item_id - 102
            db_id = int(ab_id / 2) # check
            db_item = database_ablation[db_id]
            method_type = db_item['type']
            
            if method_type not in ablation.keys():
                ablation[method_type] = {}
            if test_id not in ablation[method_type].keys():
                ablation[method_type][test_id] = {}
            if db_item['sheet'] not in ablation[method_type][test_id].keys(): 
                ablation[method_type][test_id][db_item['sheet']] = {}
            if ab_id % 2 == 0:
                ablation[method_type][test_id][db_item['sheet']]['plot'] = answer
            else:
                ablation[method_type][test_id][db_item['sheet']]['script'] = answer
            pass
        pass
    pass

pass
                
            
ab_results = {}
for key, value in ablation.items():
    if key not in ab_results.keys():
        ab_results[key] = {}
    for test_id, test_res in value.items():
        for sheet_id, sheet_res in test_res.items():
            if 'plot' not in ab_results[key].keys():
                ab_results[key]['plot'] = []
            ab_results[key]['plot'].append(sheet_res['plot'])
            if 'script' not in ab_results[key].keys():
                ab_results[key]['script'] = []
            ab_results[key]['script'].append(sheet_res['script'])
            pass
        pass
    pass

ablation_final = {}

for type, value in ab_results.items():
    for attr, res in value.items():
        mean = np.mean(res)
        std = np.std(res, ddof=1)
        n = len(res)
        conf_int_low, conf_int_high = stats.t.interval(0.95, n-1, loc=mean, scale=std/np.sqrt(n))
        print(f'{type} {attr} {mean:.2f} [{conf_int_low:.2f}, {conf_int_high:.2f}]')
        if attr not in ablation_final.keys():
            ablation_final[attr] = {}
        if type not in ablation_final[attr].keys():
            ablation_final[attr][type] = {}
        ablation_final[attr][type]['mean'] = mean
        ablation_final[attr][type]['conf_int_low'] = conf_int_low
        ablation_final[attr][type]['conf_int_high'] = conf_int_high




base_font_size = 14


plt.rcParams["font.family"] = "Times New Roman"

categories = ["Central Belief", "Motivation", "Personality", "Relationship"]
index_cat = ["central_belief", "motivation", "personality", "relationship"]
ga_values = [controllability['GA'][index]['accuracy'] for index in index_cat]
ga_errors = [(controllability['GA'][index]['accuracy'] - controllability['GA'][index]['conf_int_low'], controllability['GA'][index]['conf_int_high'] - controllability['GA'][index]['accuracy']) for index in index_cat]
psycho_values = [controllability['Psycho'][index]['accuracy'] for index in index_cat]
psycho_errors = [(controllability['Psycho'][index]['accuracy'] - controllability['Psycho'][index]['conf_int_low'], controllability['Psycho'][index]['conf_int_high'] - controllability['Psycho'][index]['accuracy']) for index in index_cat]

ga_errors_lower, ga_errors_upper = zip(*ga_errors)
ga_errors = [ga_errors_lower, ga_errors_upper]
psycho_errors_lower, psycho_errors_upper = zip(*psycho_errors)
psycho_errors = [psycho_errors_lower, psycho_errors_upper]

x = np.arange(len(categories))  
width = 0.3

fig, ax = plt.subplots(figsize=(6, 2.5))


for i in range(len(ga_values)):
    ax.bar(x[i] - width/2, ga_values[i], width, yerr=[ga_errors_lower[i:i+1], ga_errors_upper[i:i+1]],
           label='Generative Agents' if i == 0 else "", color="#cae3f7", capsize=base_font_size, error_kw=dict(capsize=2, capthick=1.5))

for i in range(len(psycho_values)):
    ax.bar(x[i] + width/2, psycho_values[i], width, yerr=[psycho_errors_lower[i:i+1], psycho_errors_upper[i:i+1]],
           label='SocioMind' if i == 0 else "", color="#d7b452", capsize=base_font_size, error_kw=dict(capsize=2, capthick=1.5))

ax.set_ylabel('Accuracy', fontsize=base_font_size)
ax.set_yticks(np.arange(0, 0.85, 0.2))
ax.set_xticks(x)
ax.set_xticklabels(categories, fontsize=base_font_size-2)
ax.legend(loc='upper left')
ax.yaxis.grid(True) 
ax.set_axisbelow(True) 

# plt.show()
plt.savefig('controllibility.pdf', format='pdf', bbox_inches='tight')
plt.close()
plt.rcParams["font.family"] = "Times New Roman"

categories = ["Generative Agents", "w\o Persona Instruction", "w\o Topic Proposal", "w\o Psychological Reflection", "Full Architecture"]

index_cat = ["GA", "instruction", "topic", "reflection", "Psycho"]


plot_values = [ablation_final['plot'][index]['mean'] for index in index_cat]
plot_errors = [(ablation_final['plot'][index]['mean'] - ablation_final['plot'][index]['conf_int_low'], ablation_final['plot'][index]['conf_int_high'] - ablation_final['plot'][index]['mean']) for index in index_cat]
script_values = [ablation_final['script'][index]['mean'] for index in index_cat]
script_errors = [(ablation_final['script'][index]['mean'] - ablation_final['script'][index]['conf_int_low'], ablation_final['script'][index]['conf_int_high'] - ablation_final['script'][index]['mean']) for index in index_cat]


# 转换误差数据格式
plot_errors_lower, plot_errors_upper = zip(*plot_errors)
plot_errors = [plot_errors_lower, plot_errors_upper]
script_errors_lower, script_errors_upper = zip(*script_errors)
script_errors = [script_errors_lower, script_errors_upper]

x = np.arange(2) 
width = 0.15
fig, ax = plt.subplots(figsize=(8, 3.5))
values = [plot_values, script_values]
errors = [plot_errors, script_errors]

colors = [ "#cae3f7", "#d6e799", "#eddf86", "#f9d547", "#d7b452",
]


categories_new = ['Plot Consistency', 'Psychological State Consistency']
for i in range(len(categories_new)):
    for j in range(len(categories)):
        ax.bar(x[i] + width * (j-2), values[i][j], width, yerr=[errors[i][0][j:j+1], errors[i][1][j:j+1]],
            label=categories[j] if i == 0 else "", color=colors[j], capsize=base_font_size, error_kw=dict(capsize=3, capthick=1.5))



ax.set_ylabel('Scores', fontsize=base_font_size+2)
ax.set_yticks(np.arange(0, 9, 2))
ax.set_ylim(1, 7.5)
# ax.set_title('Consistency on Social Evolution', fontsize=16)
ax.set_xticks(x)
ax.set_xlim(-0.6, 1.5)
ax.set_xticklabels(categories_new, fontsize=base_font_size+2)
ax.legend(loc='upper left')

ax.yaxis.grid(True) 
ax.set_axisbelow(True) 

# plt.show()
plt.savefig('consistency.pdf', format='pdf', bbox_inches='tight')
plt.close()



# # https://trueskill.org/

# from trueskill import Rating

# r1 = Rating()  # 1P's skill
# r2 = Rating()  # 2P's skill
# r3 = Rating()  # 3P's skill
# t1 = [r1]  # Team A contains just 1P
# t2 = [r2, r3]  # Team B contains 2P and 3P

# print('{:.1%} chance to draw'.format(quality([t1, t2])))
# # 13.5% chance to draw
# (new_r1,), (new_r2, new_r3) = rate([t1, t2], ranks=[0, 1])
# print(new_r1)
# # trueskill.Rating(mu=33.731, sigma=7.317)
# print(new_r2)
# # trueskill.Rating(mu=16.269, sigma=7.317)
# print(new_r3)
# # trueskill.Rating(mu=16.269, sigma=7.317)